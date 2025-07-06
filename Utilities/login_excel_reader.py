from openpyxl import load_workbook

class ExcelFunction:

    def __init__(self, file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name
        self.workbook = load_workbook(self.file)
        self.sheet = self.workbook[self.sheet]

    def row_count(self):
        return self.sheet.max_row

    def col_count(self):
        return self.sheet.max_column

    def read_data(self, row_number, col_number):
        return self.sheet.cell(row=row_number, column=col_number).value

    def write_data(self, row_number, col_number, data):
        self.sheet.cell(row=row_number, column=col_number).value = data
        self.workbook.save(self.file)